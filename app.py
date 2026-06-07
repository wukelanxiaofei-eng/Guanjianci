# -*- coding: utf-8 -*-
"""
OZON 关键词标签 & 产品简述 生成工具 (v1)

输入：本地 SKU 信息表(Excel) + OZON 热搜词导出表(Excel)
处理：调用阿里云百炼(通义千问)生成 俄语 #标签 和 俄语简述
输出：一张"待审核结果表"(Excel) + 网页内逐条预览/复制

说明：v1 全程不接 OZON API。生成结果需人工审核后，手动粘贴到 OZON 后台。
运行：在本文件所在目录执行  streamlit run app.py
"""
import os
import re
import io
import json
import zipfile
import concurrent.futures
from xml.etree import ElementTree as ET
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openai import OpenAI

# 阿里云百炼 OpenAI 兼容接口（中国大陆/北京）
BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"

# SKU 表的列名（支持多种写法，方便同事随便填）
COLS = {
    "sku":      ["货号(SKU)", "货号", "SKU", "sku"],
    "name":     ["产品名称/品类", "产品名称", "品类", "名称"],
    "size":     ["码段(尺码范围)", "码段", "尺码范围", "尺码"],
    "material": ["核心材质", "材质"],
    "audience": ["目标人群", "人群"],
    "color":    ["颜色"],
    "scene":    ["适用场景", "场景"],
    "season":   ["适用季节", "季节"],
    "image":    ["产品图(链接或路径)", "产品图", "图片链接", "图片"],
}

TEMPLATE_HEADERS = ["货号(SKU)", "产品名称/品类", "码段(尺码范围)", "核心材质",
                    "目标人群", "颜色", "适用场景", "适用季节", "产品图(链接或路径)"]
TEMPLATE_EXAMPLE = ["G6440M-3", "男士运动鞋/休闲鞋", "41-45", "弹性针织面料",
                    "男士/通勤/日常", "深灰色", "通勤 散步 旅行 运动", "春夏秋",
                    "https://example.com/g6440m-3.jpg"]


# ---------------------------------------------------------------- 配置：API Key
def load_api_key():
    """优先级：环境变量 DASHSCOPE_API_KEY > 同目录 config.json"""
    key = os.environ.get("DASHSCOPE_API_KEY", "").strip()
    if key:
        return key
    cfg = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    if os.path.exists(cfg):
        try:
            with open(cfg, "r", encoding="utf-8") as f:
                return str(json.load(f).get("api_key", "")).strip()
        except Exception:
            return ""
    return ""


# ---------------------------------------------------------------- 健壮地读 xlsx
def _local(tag: str) -> str:
    return tag.split("}")[-1]


def _col_idx(ref: str) -> int:
    letters = "".join(ch for ch in ref if ch.isalpha())
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch.upper()) - 64)
    return idx - 1


def _read_xlsx_raw(file):
    """绕过样式表，直接从 xlsx(zip) 里读第一个工作表，返回 list[list[str]]。
    用于兼容 OZON 等导出的非标准 xlsx（openpyxl 会因样式 XML 报错）。"""
    if hasattr(file, "seek"):
        file.seek(0)
    z = zipfile.ZipFile(file)
    names = z.namelist()
    shared = []
    if "xl/sharedStrings.xml" in names:
        root = ET.fromstring(z.read("xl/sharedStrings.xml"))
        for si in list(root):
            shared.append("".join(t.text or "" for t in si.iter() if _local(t.tag) == "t"))
    sheets = sorted(n for n in names if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"))
    root = ET.fromstring(z.read(sheets[0]))
    rows = []
    for el in root.iter():
        if _local(el.tag) != "row":
            continue
        cells = {}
        max_c = -1
        for c in el:
            if _local(c.tag) != "c":
                continue
            ref = c.attrib.get("r", "")
            ci = _col_idx(ref) if ref else len(cells)
            t = c.attrib.get("t", "")
            val = ""
            if t == "s":
                v = c.find("./{*}v")
                if v is not None and v.text is not None:
                    val = shared[int(v.text)]
            elif t == "inlineStr":
                val = "".join(x.text or "" for x in c.iter() if _local(x.tag) == "t")
            else:
                v = c.find("./{*}v")
                val = v.text if (v is not None and v.text is not None) else ""
            cells[ci] = val
            max_c = max(max_c, ci)
        rows.append([cells.get(i, "") for i in range(max_c + 1)])
    return rows


def _load_rows(file):
    """优先用 pandas/openpyxl；失败则用兜底解析器。返回 list[list]。"""
    try:
        if hasattr(file, "seek"):
            file.seek(0)
        raw = pd.read_excel(file, header=None, dtype=object)
        return raw.where(pd.notna(raw), "").values.tolist()
    except Exception:
        return _read_xlsx_raw(file)


# ---------------------------------------------------------------- 读取热搜词表
def parse_keywords(file) -> pd.DataFrame:
    """解析 OZON 搜索词报告，自动跳过周期/分组/排序等说明行。
    返回两列 DataFrame：query / popularity，按热度降序。"""
    rows = _load_rows(file)
    header_idx = None
    for i, r in enumerate(rows):
        if r and str(r[0]).strip() == "Запрос":
            header_idx = i
            break
    start = (header_idx + 1) if header_idx is not None else 0
    data = [(str(r[0]).strip(), r[1] if len(r) > 1 else "")
            for r in rows[start:] if r and str(r[0]).strip()]
    sub = pd.DataFrame(data, columns=["query", "popularity"])
    sub["popularity"] = pd.to_numeric(sub["popularity"], errors="coerce")
    sub = sub.dropna(subset=["popularity"])          # 自动丢掉 "—" 说明行
    sub["popularity"] = sub["popularity"].astype(int)
    sub = sub.sort_values("popularity", ascending=False).reset_index(drop=True)
    return sub


# ---------------------------------------------------------------- 读取 SKU 表
def load_skus(file) -> pd.DataFrame:
    rows = _load_rows(file)
    rows = [r for r in rows if any(str(x).strip() for x in r)]
    if not rows:
        return pd.DataFrame()
    headers = [str(h).strip() for h in rows[0]]
    body = rows[1:]
    width = len(headers)
    body = [(r + [""] * width)[:width] for r in body]
    df = pd.DataFrame(body, columns=headers).fillna("")
    return df.astype(str)


def get_field(row: dict, keys) -> str:
    for k in keys:
        if k in row and str(row[k]).strip():
            return str(row[k]).strip()
    return ""


def row_to_product(row: dict) -> dict:
    return {key: get_field(row, names) for key, names in COLS.items()}


# ---------------------------------------------------------------- 标签清洗/校验
def clean_hashtags(raw_tags, max_tags=30, max_len=30):
    """按 OZON 规则清洗标签：# 开头、词内用 _、仅字母数字下划线、单标签≤30字符、
    去重（含词序颠倒的近义重复，如 мужские_кроссовки == кроссовки_мужские）。"""
    seen = set()
    out = []
    for t in raw_tags:
        if not t:
            continue
        t = str(t).strip().lstrip("#").strip().lower()
        t = re.sub(r"[\s\-—–]+", "_", t)                  # 空格/连字符 → 下划线
        t = re.sub(r"[^a-z0-9_а-яё]", "", t)              # 只留俄/拉丁字母、数字、_
        t = re.sub(r"_+", "_", t).strip("_")
        if not t:
            continue
        if len(t) > max_len:
            t = t[:max_len].strip("_")
        key = frozenset(t.split("_"))                     # 词序无关去重
        if key in seen:
            continue
        seen.add(key)
        out.append("#" + t)
        if len(out) >= max_tags:
            break
    return out


# ---------------------------------------------------------------- 调用大模型
def _chat(client, model, messages, temperature=0.7, max_tokens=8000):
    resp = client.chat.completions.create(
        model=model, messages=messages,
        temperature=temperature, max_tokens=max_tokens,
    )
    return resp.choices[0].message.content.strip()


def _strip_json(txt: str) -> str:
    txt = txt.strip()
    txt = re.sub(r"^```(json)?", "", txt).strip()
    txt = re.sub(r"```$", "", txt).strip()
    if not txt.startswith("{"):
        m = re.search(r"\{.*\}", txt, re.S)
        if m:
            txt = m.group(0)
    return txt


def gen_tags_and_keywords(client, model, p, candidates, n_candidates=200):
    """第一步：筛同类相关热搜词 + 生成 30 个俄语标签。"""
    cand = candidates.head(n_candidates)
    cand_text = "\n".join(f"{r.query}\t{r.popularity}" for r in cand.itertuples())
    sys = ("Ты — эксперт по SEO-оптимизации карточек товаров на маркетплейсе OZON. "
           "Все ключевые слова и хештеги — строго на русском языке.")
    user = f"""Информация о товаре (вход может быть на китайском — понимай и работай на русском):
- Название/категория: {p['name']}
- Размерный ряд: {p['size']}
- Материал: {p['material']}
- Целевая аудитория: {p['audience']}
- Цвет: {p['color']}
- Сценарии использования: {p['scene']}
- Сезон: {p['season']}

Список популярных поисковых запросов покупателей OZON (запрос<TAB>популярность):
{cand_text}

Задачи:
1) Выбери из списка ТОЛЬКО запросы, реально релевантные ИМЕННО этому товару
   (верный пол, тип, материал). Сначала релевантность, потом популярность.
   Верни 15–25 запросов.
2) Сгенерируй до 30 хештегов по правилам OZON:
   - начинай с #, слова внутри хештега соединяй знаком _;
   - только буквы, цифры и _; каждый хештег не длиннее 30 символов;
   - НЕ дублируй характеристики из карточки (бренд, точный размер);
   - НЕ добавляй хештеги одинакового смысла (синонимы и перестановки слов);
   - покрой РАЗНЫЕ грани: назначение, сценарий, стиль, аудитория, сезон, материал.

Верни СТРОГО JSON без markdown и без пояснений:
{{"keywords": ["..."], "hashtags": ["#...", "..."]}}"""
    txt = _chat(client, model, [{"role": "system", "content": sys},
                                {"role": "user", "content": user}],
                temperature=0.5, max_tokens=2000)
    data = json.loads(_strip_json(txt))
    return data.get("keywords", []), data.get("hashtags", [])


def gen_description(client, model, p, keywords, min_chars=5000, max_chars=6000):
    """第二步：写 5000–6000 字符俄语简述，自然穿插所选热搜词。"""
    kw = ", ".join(keywords)
    sys = ("Ты — профессиональный копирайтер карточек товаров OZON. "
           "Пиши ТОЛЬКО на русском языке, живым маркетинговым стилем.")
    user = f"""Напиши описание товара для карточки OZON.

Информация о товаре (вход может быть на китайском — пиши результат на русском):
- Название/категория: {p['name']}
- Размерный ряд: {p['size']}
- Материал: {p['material']}
- Целевая аудитория: {p['audience']}
- Цвет: {p['color']}
- Сценарии использования: {p['scene']}
- Сезон: {p['season']}

Ключевые слова, которые нужно ЕСТЕСТВЕННО вплести в текст (не списком, не подряд,
грамотно по контексту, каждое не более 1–2 раз): {kw}

Требования:
- объём строго от {min_chars} до {max_chars} символов;
- НЕ перечисляй ключевые слова и НЕ делай переспам — OZON понижает такие карточки;
- раскрой товар всесторонне: особенности и преимущества, размерный ряд, для кого
  подходит, сценарии и места использования, сезонность, цвет;
- связный текст с абзацами, без воды и повторов.

Верни только текст описания, без заголовков и пояснений."""
    return _chat(client, model, [{"role": "system", "content": sys},
                                 {"role": "user", "content": user}],
                 temperature=0.7, max_tokens=8000)


def adjust_length(client, model, desc, min_chars=5000, max_chars=6000):
    """字符数校验：太短→请模型扩写一次；太长→在句末截断到上限内。"""
    n = len(desc)
    if n < min_chars:
        user = (f"Текущее описание ({n} символов) слишком короткое. Расширь его до "
                f"{min_chars}–{max_chars} символов: добавь больше деталей о "
                f"преимуществах, материале, ощущениях при использовании, уходе и "
                f"сочетаемости. Сохрани естественность, без переспама. "
                f"Верни только обновлённый текст.\n\nОписание:\n{desc}")
        desc = _chat(client, model, [{"role": "user", "content": user}],
                     temperature=0.7, max_tokens=8000)
        n = len(desc)
    if n > max_chars:
        cut = desc[:max_chars]
        m = max(cut.rfind("."), cut.rfind("!"), cut.rfind("?"))
        desc = cut[:m + 1] if m > min_chars else cut
    return desc


# ---------------------------------------------------------------- 密度/堆砌检查
def density_report(desc, keywords):
    low = desc.lower()
    total_words = max(len(re.findall(r"[^\W\d_]+", desc, re.UNICODE)), 1)
    counts, kw_words, flagged = [], 0, []
    for k in keywords:
        c = low.count(k.lower())
        counts.append((k, c))
        kw_words += c * len(k.split())
        if c >= 4:
            flagged.append(f"{k}×{c}")
    density = round(kw_words / total_words * 100, 1)
    return counts, density, flagged


# ---------------------------------------------------------------- 生成结果表
def build_results_excel(results) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "待审核结果"
    headers = ["货号(SKU)", "标签(空格分隔，可直接粘贴)", "标签数",
               "简述(俄语)", "简述字符数", "使用的热搜词", "关键词密度%", "风险提示"]
    ws.append(headers)
    head_font = Font(name="Arial", bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", start_color="2F5597")
    thin = Side(style="thin", color="D9D9D9")
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(thin, thin, thin, thin)
    for r in results:
        ws.append([
            r["sku"], " ".join(r["tags"]), len(r["tags"]),
            r["desc"], r["chars"], ", ".join(r["keywords"]),
            r["density"], r["risk"],
        ])
    widths = [16, 50, 8, 90, 12, 40, 12, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(name="Arial")
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_excel() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SKU信息"
    ws.append(TEMPLATE_HEADERS)
    ws.append(TEMPLATE_EXAMPLE)
    head_font = Font(name="Arial", bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", start_color="2F5597")
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in ws[2]:
        c.font = Font(name="Arial", italic=True, color="808080")
    for i in range(1, len(TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def process_one(client, model, p, kws, n_cand, min_chars, max_chars):
    """处理单个 SKU，返回结果字典（含风险提示）。供并发调用。"""
    sku = p.get("sku") or p.get("__fallback") or "未命名"
    try:
        sel_kw, raw_tags = gen_tags_and_keywords(client, model, p, kws, n_cand)
        tags = clean_hashtags(raw_tags)
        desc = gen_description(client, model, p, sel_kw, min_chars, max_chars)
        desc = adjust_length(client, model, desc, min_chars, max_chars)
        _, density, flagged = density_report(desc, sel_kw)
        risk = []
        if len(desc) < min_chars:
            risk.append(f"简述偏短({len(desc)}字符)")
        if len(tags) < 30:
            risk.append(f"标签不足30个({len(tags)})")
        if flagged:
            risk.append("疑似堆砌:" + "、".join(flagged))
        if density > 6:
            risk.append(f"关键词密度偏高({density}%)")
        return {"sku": sku, "tags": tags, "keywords": sel_kw, "desc": desc,
                "chars": len(desc), "density": density,
                "risk": "；".join(risk) if risk else "✓ 正常"}
    except Exception as e:
        return {"sku": sku, "tags": [], "keywords": [], "desc": "", "chars": 0,
                "density": 0, "risk": f"生成失败：{e}"}


# ================================================================ Streamlit 界面
st.set_page_config(page_title="OZON 标签&简述生成工具", page_icon="🛒", layout="wide")
st.title("🛒 OZON 关键词标签 & 产品简述生成工具")
st.caption("本地填表 → 自动生成俄语 #标签 和简述 → 人工审核后粘贴到 OZON 后台（v1 不接 OZON API）")

with st.sidebar:
    st.header("⚙️ 设置")
    api_key = st.text_input("阿里云百炼 API Key (sk-开头)", value=load_api_key(),
                            type="password",
                            help="也可放到环境变量 DASHSCOPE_API_KEY 或同目录 config.json，免每次输入。")
    api_key = (api_key or "").strip()
    model = st.selectbox("生成模型",
                         ["qwen-plus", "qwen-max", "qwen3-max", "qwen-flash"],
                         index=0,
                         help="简述建议 qwen-plus；追求质量用 qwen-max/qwen3-max；qwen-flash 最省但俄语长文偏弱。")
    n_cand = st.slider("送给模型筛选的候选热搜词数量", 50, 500, 200, 50,
                       help="从热搜词表里按热度取前 N 个交给模型做相关度筛选。")
    workers = st.slider("并发数（同时处理几个产品）", 1, 8, 3,
                        help="越大越快，但免费额度有调用频率上限，过大可能报错限流。建议 3–4。")
    c1, c2 = st.columns(2)
    min_chars = c1.number_input("简述最少字符", 1000, 6000, 5000, 100)
    max_chars = c2.number_input("简述最多字符", 1000, 6000, 6000, 100)
    st.divider()
    st.download_button("📄 下载 SKU 填写模板", data=build_template_excel(),
                       file_name="SKU填写模板.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

col_a, col_b = st.columns(2)
sku_file = col_a.file_uploader("① 上传 SKU 信息表（按模板填写）", type=["xlsx", "xls"])
kw_file = col_b.file_uploader("② 上传 OZON 热搜词导出表", type=["xlsx", "xls"])

run = st.button("🚀 开始生成", type="primary", use_container_width=True)

if run:
    if not api_key:
        st.error("请先在左侧填写阿里云百炼 API Key。")
        st.stop()
    if not sku_file or not kw_file:
        st.error("请同时上传 SKU 信息表 和 热搜词表。")
        st.stop()
    try:
        skus = load_skus(sku_file)
        kws = parse_keywords(kw_file)
    except Exception as e:
        st.error(f"读取表格失败：{e}")
        st.stop()

    st.info(f"已读取 {len(skus)} 个 SKU，热搜词 {len(kws)} 条。开始生成……")
    client = OpenAI(api_key=api_key, base_url=BASE_URL, timeout=300)

    products = []
    for idx, (_, row) in enumerate(skus.iterrows()):
        p = row_to_product(row.to_dict())
        p["__fallback"] = f"第{idx + 1}行"
        products.append(p)

    results = [None] * len(products)
    bar = st.progress(0.0)
    status = st.empty()
    done = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=int(workers)) as ex:
        futs = {ex.submit(process_one, client, model, p, kws, n_cand,
                          min_chars, max_chars): i for i, p in enumerate(products)}
        for fut in concurrent.futures.as_completed(futs):
            i = futs[fut]
            results[i] = fut.result()
            done += 1
            bar.progress(done / len(products))
            status.write(f"已完成 {done}/{len(products)}（{results[i]['sku']}）")

    st.session_state["results"] = results

if "results" in st.session_state:
    results = st.session_state["results"]
    st.success(f"完成，共 {len(results)} 个 SKU。请逐条审核后再使用。")
    st.download_button("⬇️ 下载全部结果（Excel，待审核）",
                       data=build_results_excel(results),
                       file_name="OZON生成结果_待审核.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary")
    st.divider()
    for r in results:
        flag = "⚠️" if r["risk"] != "✓ 正常" else "✅"
        with st.expander(f"{flag} {r['sku']} — 标签{len(r['tags'])}个 / 简述{r['chars']}字符 / {r['risk']}"):
            st.markdown("**标签（点右上角复制，直接粘贴到 OZON 的 Хештеги 字段）**")
            st.code(" ".join(r["tags"]) if r["tags"] else "（无）", language=None)
            st.markdown("**简述（俄语，点右上角复制粘贴到 Annotation 字段）**")
            st.code(r["desc"] if r["desc"] else "（无）", language=None)
            st.markdown(f"**使用的热搜词：** {', '.join(r['keywords']) or '（无）'}")
            st.caption(f"关键词密度 {r['density']}%　风险提示：{r['risk']}")
